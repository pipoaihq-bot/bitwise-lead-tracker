#!/usr/bin/env python3
"""
Pipo CSV/XLSX Import Pipeline für Bitwise Lead Tracker
Unterstützt: Apollo, LinkedIn, Notion, Custom XLSX, Gmail-Campaign-Formate
Features: Auto-Mapping, Region-Erkennung, Deduplizierung, Tier-Zuweisung
"""

import sqlite3
import csv
import os
import sys
import json
import re
from datetime import datetime
from typing import Optional, Dict, List, Tuple, Any
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DB_PATH = os.environ.get(
    "DB_PATH",
    "/Users/philippsandor/.openclaw/workspace/bitwise/leadtracker/bitwise_leads.db"
)

# ── Region-Erkennungs-Mapping ──────────────────────────────────────────────────
SWISS_KEYWORDS = ["schweiz", "swiss", "zürich", "zuerich", "zurich", "genf", "geneva",
                  "basel", "bern", "lugano", ".ch", "CH", "liechtenstein"]
UK_KEYWORDS = ["london", "uk", "england", "scotland", "wales", "british", ".co.uk",
               "city of london", "manchester", "edinburgh"]
UAE_KEYWORDS = ["dubai", "abu dhabi", "uae", "emirates", "difc", "adgm", "doha",
                "qatar", "riyadh", "saudi", "bahrain", "kuwait", "oman", "gcc"]
NORDIC_KEYWORDS = ["sweden", "schweden", "stockholm", "oslo", "norway", "norwegen",
                   "denmark", "dänemark", "finland", "finnland", "helsinki", "kopenhagen",
                   "copenhagen", ".se", ".no", ".dk", ".fi"]

# ── Kunden-Segment-Mapping für Bitwise EMEA ───────────────────────────────────
INSTITUTIONAL_KEYWORDS = [
    "bank", "versicherung", "insurance", "pension", "stiftung", "foundation",
    "asset management", "family office", "hedge fund", "private equity",
    "wealth management", "vermögen", "kapital", "fonds", "fund", "trust",
    "custodian", "depot", "verwahrstelle", "broker", "endowment"
]

def detect_region(text: str) -> str:
    """Erkennt Region aus beliebigem Text (Company, Website, Notizen)"""
    if not text:
        return "DE"
    t = text.lower()
    for kw in UAE_KEYWORDS:
        if kw.lower() in t:
            return "UAE"
    for kw in UK_KEYWORDS:
        if kw.lower() in t:
            return "UK"
    for kw in SWISS_KEYWORDS:
        if kw.lower() in t:
            return "CH"
    for kw in NORDIC_KEYWORDS:
        if kw.lower() in t:
            return "NORDICS"
    return "DE"

def detect_tier(company: str, title: str = "", aum: float = 0) -> int:
    """Weist Tier basierend auf verfügbaren Infos zu"""
    if aum >= 10_000:
        return 1  # 10B+ AUM → Tier 1
    if aum >= 1_000:
        return 2  # 1B-10B
    if aum >= 100:
        return 3  # 100M-1B

    # Fallback: Titel-basiert
    title_l = (title or "").lower()
    if any(kw in title_l for kw in ["cio", "cfo", "ceo", "managing director", "head of", "chief"]):
        return 1
    if any(kw in title_l for kw in ["director", "vp", "vice president", "partner"]):
        return 2
    if any(kw in title_l for kw in ["manager", "analyst", "senior"]):
        return 3

    # Company-basiert
    company_l = (company or "").lower()
    if any(kw in company_l for kw in ["blackrock", "vanguard", "ubs", "deutsche bank", "allianz",
                                       "axa", "generali", "zurich", "swiss re", "mnchener"]):
        return 1
    return 4  # Default

def detect_industry(company: str, title: str = "", notes: str = "") -> str:
    """Erkennt Industrie aus verfügbaren Infos"""
    combined = f"{company} {title} {notes}".lower()
    if any(kw in combined for kw in ["versicherung", "insurance", "assurance"]):
        return "Insurance"
    if any(kw in combined for kw in ["stiftung", "foundation", "endowment"]):
        return "Foundation"
    if any(kw in combined for kw in ["pension", "rentenfonds", "altersvorsorge"]):
        return "Pension Fund"
    if any(kw in combined for kw in ["family office", "single family"]):
        return "Family Office"
    if any(kw in combined for kw in ["hedge", "alternative"]):
        return "Hedge Fund"
    if any(kw in combined for kw in ["bank", "sparkasse", "landesbank", "volksbank"]):
        return "Bank"
    if any(kw in combined for kw in ["asset management", "vermögensverwaltung", "kapitalverwaltung"]):
        return "Asset Management"
    if any(kw in combined for kw in ["custodian", "verwahrstelle", "depot"]):
        return "Custodian"
    if any(kw in combined for kw in ["real estate", "immobilien", "immo"]):
        return "Real Estate"
    return "Institutional"

# ── Column-Mapper für verschiedene Formate ────────────────────────────────────

COLUMN_ALIASES = {
    "company": ["company", "firma", "unternehmen", "account name", "organisation",
                "organization", "employer", "arbeitgeber", "institution"],
    "contact_person": ["contact_person", "name", "full name", "kontakt", "ansprechpartner",
                       "contact", "person"],
    "first_name": ["first name", "firstname", "vorname", "fname"],
    "last_name": ["last name", "lastname", "nachname", "lname"],
    "title": ["title", "position", "jobtitle", "job title", "rolle", "role",
               "designation", "function", "funktion", "beruf"],
    "email": ["email", "e-mail", "mail", "emailaddress", "email address", "e_mail"],
    "phone": ["phone", "telefon", "tel", "mobile", "mobil", "nummer", "number"],
    "linkedin": ["linkedin", "linkedin url", "linkedin_url", "profil", "profile url"],
    "website": ["website", "web", "url", "homepage", "domain"],
    "region": ["region", "country", "land", "location", "standort", "city", "stadt"],
    "industry": ["industry", "branche", "sektor", "sector", "kategorie", "category",
                 "account category", "type", "account type"],
    "aum": ["aum", "aum_estimate_millions", "assets", "assets under management",
            "volume", "volumen"],
    "tier": ["tier", "priorität", "priority", "rank"],
    "notes": ["notes", "notizen", "bemerkung", "beschreibung", "description",
               "kommentar", "comment", "status"],
}

def normalize_col(col: str) -> str:
    """Normalisiert Spaltennamen für Vergleiche"""
    return re.sub(r'[^a-z0-9]', '', col.lower().strip())

def map_columns(headers: List[str]) -> Dict[str, int]:
    """Mappt Spaltennamen auf interne Felder. Gibt {field: col_index} zurück"""
    norm_headers = [(i, normalize_col(h)) for i, h in enumerate(headers) if h]
    mapping = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = normalize_col(alias)
            for i, norm_h in norm_headers:
                if norm_h == norm_alias and field not in mapping:
                    mapping[field] = i
                    break
    return mapping

def get_cell(row: List[Any], col_map: Dict[str, int], field: str, default: str = "") -> str:
    """Holt Zellwert sicher aus Row"""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return str(val).strip() if val is not None else default

# ── Deduplizierung ─────────────────────────────────────────────────────────────

def get_existing_fingerprints(conn: sqlite3.Connection) -> set:
    """Holt alle existierenden Fingerprints (email + company_norm)"""
    rows = conn.execute("SELECT email, company FROM leads").fetchall()
    fps = set()
    for row in rows:
        if row[0]:
            fps.add(normalize_col(row[0]))
        if row[1]:
            fps.add(normalize_col(row[1]))
    return fps

def build_fingerprint(email: str, company: str) -> List[str]:
    """Erstellt Fingerprints für einen Lead"""
    fps = []
    if email and email != "n/a" and "@" in email:
        fps.append(normalize_col(email))
    if company:
        fps.append(normalize_col(company))
    return fps

# ── Import-Logik ──────────────────────────────────────────────────────────────

def import_rows(
    rows: List[List[Any]],
    headers: List[str],
    source_tag: str,
    db_path: str = DB_PATH,
    dry_run: bool = False
) -> Dict[str, int]:
    """
    Importiert Rows in die Datenbank.
    Returns: {'imported': n, 'skipped_dup': n, 'skipped_invalid': n, 'errors': n}
    """
    col_map = map_columns(headers)
    print(f"\n📋 Column Mapping für '{source_tag}':")
    for field, idx in col_map.items():
        print(f"   {field} → '{headers[idx]}'")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    existing_fps = get_existing_fingerprints(conn)
    stats = {"imported": 0, "skipped_dup": 0, "skipped_invalid": 0, "errors": 0}

    for i, row in enumerate(rows):
        try:
            # Felder extrahieren
            company = get_cell(row, col_map, "company")
            first_name = get_cell(row, col_map, "first_name")
            last_name = get_cell(row, col_map, "last_name")
            contact_person = get_cell(row, col_map, "contact_person")

            # Name zusammenbauen falls nötig
            if not contact_person and (first_name or last_name):
                contact_person = f"{first_name} {last_name}".strip()

            email = get_cell(row, col_map, "email")
            title = get_cell(row, col_map, "title")
            phone = get_cell(row, col_map, "phone")
            linkedin = get_cell(row, col_map, "linkedin")
            website = get_cell(row, col_map, "website")
            region_hint = get_cell(row, col_map, "region")
            industry_hint = get_cell(row, col_map, "industry")
            notes = get_cell(row, col_map, "notes")
            tier_raw = get_cell(row, col_map, "tier")

            # Validation
            if not company or company.lower() in ["", "n/a", "none", "unknown"]:
                stats["skipped_invalid"] += 1
                continue

            # Deduplizierung
            fps = build_fingerprint(email, company)
            is_dup = any(fp in existing_fps for fp in fps)
            if is_dup:
                stats["skipped_dup"] += 1
                continue

            # Region erkennen
            region_text = f"{company} {region_hint} {website} {notes}"
            region = detect_region(region_text)
            # Normalize region to valid enum values
            region_map = {"DE": "DE", "CH": "CH", "UK": "UK", "UAE": "UAE", "NORDICS": "NORDICS"}
            region = region_map.get(region, "DE")

            # Tier bestimmen
            tier = 4
            if tier_raw:
                try:
                    tier = int(str(tier_raw).replace("TIER_", "").strip())
                    tier = max(1, min(4, tier))
                except:
                    tier = detect_tier(company, title)
            else:
                tier = detect_tier(company, title)

            # Industrie
            industry = detect_industry(company, title, f"{industry_hint} {notes}")

            # Notizen als use_case nutzen
            use_case_text = f"ETH Staking | Source: {source_tag}"
            if notes:
                use_case_text = f"{notes[:200]} | Source: {source_tag}"

            # LinkedIn validieren
            if linkedin and not linkedin.startswith("http"):
                linkedin = None

            if not dry_run:
                conn.execute("""
                    INSERT INTO leads (
                        company, region, tier, contact_person, title, email, linkedin,
                        stage, industry, use_case, staking_readiness,
                        aum_estimate_millions, expected_deal_size_millions, expected_yield
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    company, region, tier, contact_person, title,
                    email if email and "@" in email else None,
                    linkedin, "prospecting", industry, use_case_text,
                    "Unknown",
                    0.0, 0.0, 0.0
                ))

            # Fingerprints aktualisieren für diese Session
            for fp in fps:
                existing_fps.add(fp)

            stats["imported"] += 1

            if stats["imported"] % 500 == 0:
                print(f"   ⏳ {stats['imported']} importiert...")

        except Exception as e:
            stats["errors"] += 1
            print(f"   ⚠️ Row {i}: {e}")

    if not dry_run:
        conn.commit()
    conn.close()

    return stats

# ── File Loader ────────────────────────────────────────────────────────────────

def load_csv(filepath: str) -> Tuple[List[str], List[List[Any]]]:
    """Lädt CSV und gibt (headers, rows) zurück"""
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, errors="replace") as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)
                delimiters = [",", ";", "\t", "|"]
                counts = {d: sample.count(d) for d in delimiters}
                delimiter = max(counts, key=counts.get)

                reader = csv.reader(f, delimiter=delimiter)
                all_rows = [row for row in reader if any(cell.strip() for cell in row)]
                if not all_rows:
                    return [], []
                headers = all_rows[0]
                rows = all_rows[1:]
                return headers, rows
        except Exception:
            continue
    raise ValueError(f"Could not read {filepath}")

def load_xlsx(filepath: str, sheet_name: str = None) -> Tuple[List[str], List[List[Any]]]:
    """Lädt XLSX Sheet und gibt (headers, rows) zurück"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Wähle Sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Finde das Sheet mit den meisten Daten
        best = None
        best_rows = 0
        for name in wb.sheetnames:
            if any(kw in name.lower() for kw in ["kontakt", "contact", "lead", "alle", "all"]):
                ws_test = wb[name]
                if ws_test.max_row and ws_test.max_row > best_rows:
                    best = name
                    best_rows = ws_test.max_row
        if not best:
            best = wb.sheetnames[0]
        ws = wb[best]

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            all_rows.append([cell for cell in row])

    wb.close()

    if not all_rows:
        return [], []

    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    rows = [[str(c) if c is not None else "" for c in row] for row in all_rows[1:]]
    return headers, rows

# ── Haupt-Import-Funktion ──────────────────────────────────────────────────────

def import_file(
    filepath: str,
    source_tag: str = None,
    dry_run: bool = False,
    sheet_name: str = None,
    db_path: str = DB_PATH
) -> Dict[str, int]:
    """
    Importiert eine einzelne CSV oder XLSX Datei.

    Args:
        filepath: Pfad zur Datei
        source_tag: Optionaler Tag (z.B. "Apollo", "LinkedIn", "Manual")
        dry_run: Nur analysieren, nicht schreiben
        sheet_name: Spezifisches XLSX Sheet (optional)
        db_path: DB Pfad

    Returns: Import-Statistiken
    """
    filepath = str(filepath)
    ext = Path(filepath).suffix.lower()
    name = Path(filepath).name
    source = source_tag or name

    print(f"\n{'─'*60}")
    print(f"📂 Importiere: {name}")
    print(f"   Quelle: {source}")
    print(f"   {'DRY RUN - ' if dry_run else ''}DB: {db_path}")

    if ext == ".csv":
        headers, rows = load_csv(filepath)
    elif ext in [".xlsx", ".xls"]:
        headers, rows = load_xlsx(filepath, sheet_name)
    else:
        raise ValueError(f"Unbekanntes Dateiformat: {ext}")

    print(f"   Zeilen: {len(rows)}, Spalten: {len(headers)}")
    print(f"   Header: {headers[:8]}")

    if not rows:
        print("   ⚠️ Keine Daten!")
        return {"imported": 0, "skipped_dup": 0, "skipped_invalid": 0, "errors": 0}

    stats = import_rows(rows, headers, source, db_path=db_path, dry_run=dry_run)

    print(f"\n{'='*60}")
    print(f"✅ Import '{name}' abgeschlossen:")
    print(f"   ✓ Importiert:       {stats['imported']}")
    print(f"   ⟳ Duplikate:        {stats['skipped_dup']}")
    print(f"   ✗ Ungültig:         {stats['skipped_invalid']}")
    print(f"   ⚠ Fehler:           {stats['errors']}")
    print(f"{'='*60}")

    return stats

def import_directory(
    directory: str,
    source_tag: str = None,
    dry_run: bool = False,
    db_path: str = DB_PATH
) -> Dict[str, int]:
    """Importiert alle CSV/XLSX aus einem Verzeichnis"""
    total = {"imported": 0, "skipped_dup": 0, "skipped_invalid": 0, "errors": 0}
    path = Path(directory)

    files = list(path.glob("*.csv")) + list(path.glob("*.xlsx"))
    if not files:
        print(f"⚠️ Keine CSV/XLSX Dateien in {directory}")
        return total

    print(f"\n🗂️  Verzeichnis-Import: {directory}")
    print(f"   {len(files)} Dateien gefunden")

    for f in sorted(files):
        try:
            stats = import_file(str(f), source_tag=source_tag, dry_run=dry_run, db_path=db_path)
            for k in total:
                total[k] += stats[k]
        except Exception as e:
            print(f"⚠️ Fehler bei {f.name}: {e}")
            total["errors"] += 1

    print(f"\n{'='*60}")
    print(f"📊 GESAMT-ERGEBNIS ({len(files)} Dateien):")
    print(f"   ✓ Importiert:       {total['imported']}")
    print(f"   ⟳ Duplikate:        {total['skipped_dup']}")
    print(f"   ✗ Ungültig:         {total['skipped_invalid']}")
    print(f"   ⚠ Fehler:           {total['errors']}")
    print(f"{'='*60}")
    return total

def get_db_stats(db_path: str = DB_PATH) -> str:
    """Gibt aktuelle DB-Statistiken zurück"""
    conn = sqlite3.connect(db_path)
    total = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
    by_region = conn.execute(
        "SELECT region, COUNT(*) FROM leads GROUP BY region ORDER BY COUNT(*) DESC"
    ).fetchall()
    by_stage = conn.execute(
        "SELECT stage, COUNT(*) FROM leads WHERE stage != 'prospecting' GROUP BY stage"
    ).fetchall()
    conn.close()

    lines = [f"📊 Leads in DB: {total}"]
    lines.append("   By Region: " + ", ".join(f"{r}: {c}" for r, c in by_region))
    if by_stage:
        lines.append("   Active: " + ", ".join(f"{s}: {c}" for s, c in by_stage))
    return "\n".join(lines)

# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Pipo CSV/XLSX Import Pipeline für Bitwise Lead Tracker"
    )
    parser.add_argument("path", nargs="?", help="Datei oder Verzeichnis")
    parser.add_argument("--source", "-s", help="Source Tag (z.B. Apollo, LinkedIn)")
    parser.add_argument("--dry-run", "-n", action="store_true", help="Nur analysieren")
    parser.add_argument("--db", default=DB_PATH, help="DB Pfad")
    parser.add_argument("--stats", action="store_true", help="Nur DB-Stats anzeigen")
    parser.add_argument("--sheet", help="XLSX Sheet-Name")

    args = parser.parse_args()

    if args.stats:
        print(get_db_stats(args.db))
        sys.exit(0)

    if not args.path:
        print("Usage: python3 csv_importer.py <file_or_dir> [--source Apollo] [--dry-run]")
        print("\nBeispiele:")
        print("  python3 csv_importer.py ~/Downloads/apollo_export.csv --source Apollo")
        print("  python3 csv_importer.py ~/Downloads/ --source LinkedIn --dry-run")
        print("  python3 csv_importer.py --stats")
        sys.exit(1)

    p = Path(args.path)
    if p.is_dir():
        import_directory(str(p), source_tag=args.source, dry_run=args.dry_run, db_path=args.db)
    elif p.is_file():
        import_file(str(p), source_tag=args.source, dry_run=args.dry_run,
                    sheet_name=args.sheet, db_path=args.db)
    else:
        print(f"❌ Pfad nicht gefunden: {args.path}")
        sys.exit(1)

    print(f"\n{get_db_stats(args.db)}")
